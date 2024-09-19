from tkinter import *
from tkinter import ttk
import datetime
from tkinter import messagebox
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime, timedelta
from dateutil.relativedelta import *
from tkcalendar import *

class MyCalendar(Calendar):


    def _next_month(self):
        Calendar._next_month(self)
        self.event_generate('<<CalendarMonthChanged>>')

    def _prev_month(self):
        Calendar._prev_month(self)
        self.event_generate('<<CalendarMonthChanged>>')

    def _next_year(self):
        Calendar._next_year(self)
        self.event_generate('<<CalendarMonthChanged>>')

    def _prev_year(self):
        Calendar._prev_year(self)
        self.event_generate('<<CalendarMonthChanged>>')

    def get_displayed_month_year(self):
        return self._date.month, cal._date.year

def inputInfo():

    wb = openpyxl.load_workbook('creditSheet.xlsx')

        
    card_select = ccOption.get()


    if card_select not in wb.sheetnames:
            
        messagebox.askretrycancel('askretrycancel', " '{card}' is not an option. Please select a Card Option".format(card = card_select))
        return


    currentSheet = wb[card_select]

    newRow = str(currentSheet.max_row + 1)

    inputValues = []
    

    transType = transactionBox.get()

    if transType == '':
        messagebox.askretrycancel('askretrycancel', " Please select a Transaction Type")
        return

    inputValues.append(transType)



    transactionName = transactionName_input.get()
    inputValues.append(transactionName)

    currentDate = dateEntry.get()
    inputValues.append(currentDate)

    while True:

        money = amountEntry.get()
        
        try:

            money = float(money)
            inputValues.append(money)
            break
            
        except ValueError:

            
            messagebox.askretrycancel('askretrycancel', 'Type in a number')
            return          
    

    for i in range(currentSheet.max_column):
        newCell = get_column_letter(i+1) + newRow

        currentSheet[newCell] = inputValues[i]

    wb.save('creditSheet.xlsx')


    amountEntry.delete(0, '')
    transactionName_input.delete(0, '')
    transactionBox.set('')
    ccOption.set('Choose Credit Card')

def show_trans():

    wb = openpyxl.load_workbook('creditSheet.xlsx')
    dt = datetime.now()
    empty_transaction_label.config(text = '')

    for item in ccExpense.get_children():
        ccExpense.delete(item)
    
    cc_card = cc_options.get()

    currentSheet = wb[cc_card]

    card_due_sheet = wb['Card Due Date']

    for cards in range(card_due_sheet.max_row):

        current_cell = 'A'+ str(cards+1)
        
        if card_due_sheet[current_cell].value == cc_card:
            
            date_cell = 'B' + str(cards+1)
            due_date = str(card_due_sheet[date_cell].value)

            #act_due_date.config(text = due_date)
            dueMonth = str(dt.month)
            dueYear = str(dt.year)

            complete_due_date = dueYear+'/'+dueMonth+'/'+due_date
            complete_stripped = datetime.strptime(complete_due_date, '%Y/%m/%d')


            max_row = currentSheet.max_row
            working_cell = 'A' + str(max_row)

            while currentSheet[working_cell].value not in ('Payed', 'Transaction Type'):

                max_row -= 1
                working_cell = 'A' + str(max_row)
                

            if max_row == 1:

                payment_date = 'C' + str(2)

            else:

                payment_date = 'C' + str(max_row)


            old_payment_date = currentSheet[payment_date].value

            if old_payment_date is None:

                empty_transaction_label.config(text = 'NO TRANSACTIONS!')

                if  complete_stripped < dt:

                    due_date = complete_stripped + relativedelta(months =+ 1)

                    days_remaining = due_date - dt
                    days_remaining_input.config(text = '{days}'.format(days = days_remaining.days))

                    due_date = due_date.strftime('%m/%d')
                    act_due_date.config(text = due_date)

                    card_balance_total.config(text = '0.00')

                    

                else:

                    days_remaining = complete_stripped - dt
                    days_remaining_input.config(text = '{days}'.format(days = days_remaining.days))
    
                    due_date = datetime.strftime(complete_stripped, '%m/%d')
                    act_due_date.config(text = due_date)
                    card_balance_total.config(text = '0.00')

                return
    
            else:
            
                old_payment_date = datetime.strptime(old_payment_date,'%Y/%m/%d')
            

            month_back = complete_stripped - relativedelta(months =+ 1)

            if month_back < old_payment_date < dt:


                due_date = complete_stripped + relativedelta(months =+ 1)

                days_remaining = due_date - dt
                days_remaining_input.config(text = '{days}'.format(days = days_remaining.days))

                due_date = due_date.strftime('%m/%d')
                act_due_date.config(text = due_date)

                over_due_input.config(text = '')

            else:

                due_date = datetime.strftime(complete_stripped, '%m/%d')

                new_due_date = complete_stripped + relativedelta(months=+1)
                days_remaining = new_due_date - dt
                
                days_over_due = dt - complete_stripped

                new_due_date = new_due_date.strftime('%m/%d')
                    
                act_due_date.config(text='Past Due Date: {old_date} \nNew Due Date: {date}'.format(old_date=due_date ,date=new_due_date))
                over_due_input.config(text = '{days}'.format(days=days_over_due.days))
                days_remaining_input.config(text = '{days}'.format(days = days_remaining.days))
                     
        else:
            continue  
    

    headings = ['transaction name', 'date', 'amount']

    rowValues = []

    limit = currentSheet.max_row

    card_balance = 0

    ccExpense.tag_configure('odd', background = 'white')
    ccExpense.tag_configure('even', background = 'lightblue')

    count = 0

    for i in range(limit-1):

        currentRow = str(i+2)
    

        for data in range(currentSheet.max_column):
            newCurrentCell = get_column_letter(data+1) + currentRow
            rowValues.append(currentSheet[newCurrentCell].value)

        typeCell = 'A' + str(i+2)
        amountCell = 'D' + str(i+2)


        if currentSheet[typeCell].value == 'Payed':

            total = currentSheet[amountCell].value * -1
            card_balance += total
                    
        else:

            total = (currentSheet[amountCell].value) * 1
            card_balance += total

        card_balance = round(card_balance, 2)

        if count % 2 == 0:

            ccExpense.insert(parent='', index='end', text=rowValues[0],iid=i, values = rowValues[1:], tags=('even', ))

        else:

            ccExpense.insert(parent='', index='end', text=rowValues[0],iid=i, values = rowValues[1:], tags=('odd', ))

        count += 1

        rowValues.clear()

    card_balance_total.config(text = card_balance)
    
    
        
def creditCardSub():

    wb = openpyxl.load_workbook('creditSheet.xlsx')

    cc_name = addEntry.get()
    due_date = addDueEntry.get()

    if cc_name == '':
        messagebox.askretrycancel('askretrycancel', 'Please type in a Credit Card Name')
        return

    if due_date == '':
        messagebox.askretrycancel('askretrycancel', 'Please select a Due Date')
        return

    if cc_name not in wb.sheetnames:

        wb.create_sheet(title = cc_name)
        wb.save('creditSheet.xlsx')

        wb = openpyxl.load_workbook('creditSheet.xlsx')

        sheet_headers = ['Transaction Type','Transaction Name', 'Date', 'Amount']
        currentSheet = wb[cc_name]

        bold24Font = Font(name='Times New Roman', size = 24, bold = True)

        for y in range(len(sheet_headers)):

            cellName = get_column_letter(y+1) + '1'
            letter = get_column_letter(y+1)

            currentSheet[cellName].font = bold24Font
            currentSheet[cellName] = sheet_headers[y]

            currentSheet.column_dimensions[letter].width = 40

        currentSheet = wb['Card Due Date']

        cc_name_cell = 'A' + str(currentSheet.max_row + 1)
        cc_due_date = 'B' + str(currentSheet.max_row + 1)

        currentSheet[cc_name_cell] = cc_name
        currentSheet[cc_due_date] = due_date


        optList.append(cc_name)

        credit_card_options['values'] = wb.sheetnames
        cc_options['values'] = wb.sheetnames
        ccOption['values'] = wb.sheetnames
        
    else:

        messagebox.showinfo('Notification', 'Credit Card Already In System')

    wb.save('creditSheet.xlsx')

    month, year = cal.get_displayed_month_year()

    due_date = str(month)+'/'+str(due_date)+'/'+str(year)
    due_date = datetime.strptime(due_date, '%m/%d/%Y')

    cal.calevent_create(due_date, cc_name, 'message')
    

    addEntry.delete(0, '')
    addDueEntry.delete(0,'')
    messagebox.showinfo('Notification', 'Submitted Successfully')

def creditCardDel():

    wb = openpyxl.load_workbook('creditSheet.xlsx')

    selection = credit_card_options.get()

    if selection == '':

        messagebox.askretrycancel('askretrycancel', 'Please select a Credit Card to delete')
        return

    del wb[selection]

    credit_card_options['values'] = wb.sheetnames
    cc_options['values'] = wb.sheetnames
    ccOption['values'] = wb.sheetnames

    card_due_sheet = wb['Card Due Date']

    for r in range(card_due_sheet.max_row):

        current_cell = 'A' + str(r+1)

        if card_due_sheet[current_cell].value != selection:

            continue
        else:

            cell_row = r + 1
            break

    card_due_sheet.delete_rows(cell_row,1)
    

    wb.save('creditSheet.xlsx')

    cal.calevent_remove('all')
    populate_calendar()
    

    credit_card_options.delete(0,'')

    messagebox.showinfo('Notification', 'Deleted Successfully')

def show_calendar(event):

    cal.calevent_remove('all')

    wb = openpyxl.load_workbook('CreditSheet.xlsx')

    current_sheet = wb['Card Due Date']

    month, year = cal.get_displayed_month_year()


    for i in range(current_sheet.max_row-1):
        event_desc_cell = 'A' + str(i+2)
        event_day_cell = 'B' + str(i+2)

        event_date_full = str(month) +'/'+ str(current_sheet[event_day_cell].value) +'/'+ str(year)
        event_date_full = datetime.strptime(event_date_full, '%m/%d/%Y')

        cal.calevent_create(event_date_full, current_sheet[event_desc_cell].value , 'message')


def populate_calendar():

    wb = openpyxl.load_workbook('CreditSheet.xlsx')

    current_sheet = wb['Card Due Date']

    month, year = cal.get_displayed_month_year()

    for i in range(current_sheet.max_row-1):
        event_desc_cell = 'A' + str(i+2)
        event_day_cell = 'B' + str(i+2)

        event_date_full = str(month) +'/'+ str(current_sheet[event_day_cell].value) +'/'+ str(year)
        event_date_full = datetime.strptime(event_date_full, '%m/%d/%Y')

        cal.calevent_create(event_date_full, current_sheet[event_desc_cell].value , 'message')


def windowCreation():

    mainWindow = Tk()
    mainWindow.geometry('1200x400')
    mainWindow.title('Credit Card Tracker')

    main_notebook = ttk.Notebook(mainWindow)
    main_notebook.pack(expand = 1, fill = BOTH)

    global addEntry
    global optList
    global credit_card_options
    global addEntry
    global wb
    global cc_options
    global ccExpense
    global transactionName_input
    global dateEntry
    global amountEntry
    global transactionBox
    global ccOption
    global card_balance_total
    global due_date_label
    global act_due_date
    global over_due_input
    global days_remaining_input
    global addDueEntry
    global empty_transaction_label
    global cal
    
    
    wb = openpyxl.load_workbook('creditSheet.xlsx')

    optList = wb.sheetnames[1:]
    days = list(range(1,32))

    style = ttk.Style(mainWindow)
    style.theme_use('clam')



    # ------------ MAIN TAB ------------#

    
    mainTab = ttk.Frame(main_notebook)
    main_notebook.add(mainTab, text = 'Main')

    mainTab.columnconfigure(3, weight=2)
    mainTab.rowconfigure(1, weight = 1)

    addLabel = ttk.Label(mainTab, text = 'Add Credit Card', font= ('caliber', 10, 'bold'))
    addLabel.grid(row = 0, column = 0)

    addEntry = ttk.Entry(mainTab, font = ('caliber', 10, 'normal'))
    addEntry.grid(row = 0, column = 1)

    add_cc_button = ttk.Button(mainTab, text = 'Submit', command = creditCardSub)
    add_cc_button.grid(row = 0, column = 2)

    delLabel = ttk.Label(mainTab, text = 'Delete Credit Card', font= ('caliber', 10, 'bold'))
    delLabel.grid(row = 2, column = 0)

    credit_card_options = ttk.Combobox(mainTab, values = optList)
    credit_card_options.set('Pick Credit Card')
    credit_card_options.grid(row = 2, column = 1)

    delete_cc_button = ttk.Button(mainTab, text = 'Delete', command = creditCardDel)
    delete_cc_button.grid(row = 2, column = 2)

    addDue = ttk.Label(mainTab, text = 'Add Credit Card Due Day', font= ('caliber', 10, 'bold'))
    addDue.grid(row = 1, column = 0)

    addDueEntry = ttk.Combobox(mainTab, values = days)
    addDueEntry.grid(row = 1, column = 1)

    cal = MyCalendar(mainTab,
                     setmode = 'day',
                     date_pattern = 'd/m/yy',
                     showothermonthdays = False,
                     background = 'black',
                     showweeknumbers = False,
                     firstweekday = 'sunday',
                     weekendforeground = 'red',
                     foreground = 'white',
                     bordercolor = 'black',
                     selectmode = 'none',
                     tooltipforeground = 'gray',
                     cursor = 'hand2',
                     tooltipdelay = 1,
                     headersbackground = 'gray',
                     headserforeground = 'white')

    cal.grid( row = 0, column = 3, ipadx = 10, ipady = 10, sticky = 'NSEW', rowspan = 3)

    cal.bind('<<CalendarMonthChanged>>', show_calendar)
    populate_calendar()



    # ------------ VIEW TAB -----------#
    
    viewTab = ttk.Frame(main_notebook)
    main_notebook.add(viewTab, text = 'View')


    cc_options = ttk.Combobox(viewTab, values = optList)
    cc_options.set('Pick Credit Card')
    cc_options.grid(row = 0, column = 0)

    chooseButton = ttk.Button(viewTab, text = 'Enter', command= show_trans)
    chooseButton.grid(row = 0, column = 1)

    ccExpense = ttk.Treeview(viewTab)
    ccExpense.grid(row = 0, column = 2)

    card_balance_label = ttk.Label(viewTab, text = 'Card Balance: $', font= ('caliber', 20, 'bold'))
    card_balance_label.grid(row = 0, column = 3)

    card_balance_total = ttk.Label(viewTab, text = '0.00', font = ('caliber', 20, 'bold'))
    card_balance_total.grid(row = 0, column = 4)

    due_date_label = ttk.Label(viewTab, text = 'Due Date: ', font = ('caliber', 20, 'bold'))
    due_date_label.grid(row = 1, column = 0)

    over_due_label = ttk.Label(viewTab, text = 'Days Over Due: ', font = ('caliber', 20, 'bold'))
    over_due_label.grid(row = 2, column = 0)

    over_due_input = ttk.Label(viewTab, textvariable = '', font = ('caliber', 20, 'bold'))
    over_due_input.grid(row = 2, column = 1)

    act_due_date = ttk.Label(viewTab, textvariable = '', font = ('caliber', 20, 'bold'))
    act_due_date.grid(row = 1, column = 1)

    days_remaining_label = ttk.Label(viewTab, text = 'Days Remaining: ', font = ('caliber', 20, 'bold'))
    days_remaining_label.grid(row = 3, column = 0)

    days_remaining_input = ttk.Label(viewTab, textvariable = '', font = ('caliber', 20, 'bold'))
    days_remaining_input.grid(row = 3, column = 1)

    empty_transaction_label = ttk.Label(viewTab, textvariable = '', font = ('caliber', 20, 'bold'))
    empty_transaction_label.grid(row = 2, column = 2)
    

    ccExpense['columns'] = ('transaction name', 'date', 'amount')

    # Formatting Columns
    
    ccExpense.column('#0', width = 120, minwidth=25)
    ccExpense.column('transaction name', anchor=CENTER, width = 120)
    ccExpense.column('date', anchor=CENTER, width = 120)
    ccExpense.column('amount', anchor=CENTER, width = 120)

    # Creating Headings

    ccExpense.heading('#0', text = "Transaction Type")
    ccExpense.heading('transaction name', text = "Transaction Name")
    ccExpense.heading('date', text = "Date")
    ccExpense.heading('amount', text = "Amount ($)")
    

    # ----------- PAY TAB ------------#

    transactionTypes = ['Spent','Payed']

    payTab = ttk.Frame(main_notebook)
    main_notebook.add(payTab, text = 'Transactions')

    ccOption = ttk.Combobox(payTab, values = optList)
    ccOption.set('Pick Credit Card')
    ccOption.grid(row = 0, column = 0)


    transaction_name_label = ttk.Label(payTab, text='Transaction Type', font = ('caliber', 10, 'bold'))
    transaction_name_label.grid(row = 1, column = 0, padx = 20, pady = 20)

    transactionBox = ttk.Combobox(payTab, values = transactionTypes)
    transactionBox.grid(row = 1, column = 1)

    amountLabel = ttk.Label(payTab, text = 'Amount ($)', font = ('caliber', 10, 'bold'))
    amountLabel.grid(row = 4, column = 0, padx = 20, pady = 20) #row 4 column 0

    amountEntry = ttk.Entry(payTab, font = ('caliber', 10, 'normal'))
    amountEntry.grid(row = 4, column = 1) #row 4 column 1

    currentDate = datetime.now()
    dateFormat = currentDate.strftime('%Y/%m/%d')

    dateLabel = ttk.Label(payTab, text = 'Date', font = ('caliber', 10, 'bold'))
    dateLabel.grid(row = 3, column = 0, pady = 20, padx = 20)

    dateEntry = ttk.Entry(payTab, font = ('caliber', 10, 'normal'))
    dateEntry.insert(0, dateFormat)
    dateEntry.grid(row = 3, column = 1)

    transactionName = ttk.Label(payTab, text='Transaction Name', font = ('caliber', 10, 'bold'))
    transactionName.grid(row = 2, column = 0, padx = 20, pady = 20)

    transactionName_input = ttk.Entry(payTab, font = ('caliber', 10, 'normal'))
    transactionName_input.grid(row = 2 , column = 1)

    submitButton = ttk.Button(payTab, text = 'Submit', command = inputInfo)
    submitButton.grid(row = 5, column = 0)

    

    wb.save('creditSheet.xlsx')
    mainWindow.mainloop()


windowCreation()
